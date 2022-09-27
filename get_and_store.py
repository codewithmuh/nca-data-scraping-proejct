
import json
import pandas as pd
from openpyxl import Workbook,load_workbook

import os 
a= os.path.isfile(f'.//Sport Men.xlsx')
if a is False:
    book=Workbook()
    sheet=book.active
else:
    book = load_workbook(f'.//Sport Men.xlsx')
    sheet = book.active


sheet.append(['Year','ncaaId','First Name','Last Name','Full Name',"Email",'Initiated Date','Last Updated','D','institutionName','sport','Sport Conference','status','transfer_id','email'])

with open('data.json', 'r') as j:
    contents = json.loads(j.read())
    for each in contents:
        Year = each.get("academicYearString")
        ncaaId = each.get("ncaaId")
        firstName = each.get("firstName")
        lastName = each.get("lastName")
        fullname = firstName+' '+lastName
        
        Initiated_Date = each.get("createTimestamp").split('T')[0]
        Last_Updated = each.get("updateTimestamp").split('T')[0]
        D = each.get("division")
        D = 'I'*int(D)
        institutionName = each.get("institutionName")
        sport = each.get("saTransferSports")[0]['description']
        Sport_Conference = each.get("saTransferSports")[0].get('confName')
        status = each.get("statusCode")
        transfer_id = each.get("transferId")
        sheet.append([Year,ncaaId,firstName,lastName,fullname,'',Initiated_Date,Last_Updated,D,institutionName,sport,Sport_Conference,status,transfer_id,''])

    book.save(f".//Sport Men.xlsx")

    print("all done file saved dear")


        