import pandas as pd
import scrapy
from openpyxl import Workbook,load_workbook
from scrapy.crawler import CrawlerProcess
import json
import os
from scrapy import signals
from pydispatch import dispatcher


if os.stat("cookie2.json").st_size > 2:
        file =  open('./cookie2.json', 'r') 
        cookiehai = json.load(file)
       



class EmailsSpider(scrapy.Spider):
    name = 'see'

    custom_settings = {
            "FEED_EXPORT_ENCODING": "utf-8-sig",

            # enable the middleware
            # 'DOWNLOADER_MIDDLEWARES' : {'scrapy_zyte_smartproxy.ZyteSmartProxyMiddleware': 610},
         
            # 'ZYTE_SMARTPROXY_ENABLED' : True, 
         
            # 'ZYTE_SMARTPROXY_APIKEY' : 'c477bd5fa667490389d951eb531a287d',
             "DOWNLOAD_DELAY": 1.5,

        
                                                                                             }



    def __init__(self):
        dispatcher.connect(self.spider_closed, signals.spider_closed)
         

    def start_requests(self):

        a= os.path.isfile(f'.//final sport men.xlsx')
        if a is False:
            self.book=Workbook()
            self.sheet=self.book.active

        else:
            self.book = load_workbook(f'.//final sport men.xlsx')
            self.sheet = self.book.active
            
        self.sheet.append(['Year','ncaaId','First Name','Last Name','Full Name',"Email",'Initiated Date','Last Updated','D','institutionName','sport','Sport Conference','status','transfer_id','email'])

            #input file read 
        df = load_workbook(f'./Sport Men.xlsx')
        ws = df['Sheet']
        for row in ws.iter_rows():
            row =[row[i].value for i in range(len(row)-1)]

            notice_id = row[-1]

            coookies = {
            'authority': 'web1.ncaa.org',
            'method': 'GET',
            'path': f'/saTransfer/notice?id={notice_id}',
            'scheme': 'https',
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'accept-encoding': 'gzip, deflate, br',
            'accept-language': 'en-US,en;q=0.9',
            'cache-control':' max-age=0',
            # 'sec-ch-ua': "Google Chrome";v="105", "Not)A;Brand";v="8", "Chromium";v="105" ,
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': "Windows",
            'sec-fetch-dest': 'document',
            'sec-fetch-mode': 'navigate',
            'sec-fetch-site': None,
            'sec-fetch-user': '?1',
            'upgrade-insecure-requests': 1,
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/105.0.0.0 Safari/537.36'

            }

            yield scrapy.Request(
                url =f'https://web1.ncaa.org/saTransfer/notice?id={notice_id}' ,
                callback=self.parse,
                headers=coookies,
                cookies=cookiehai,
                meta ={
                    'complete_row':row
                }
            )


    def parse(self, response, **kwargs):
        row_is = response.request.meta['complete_row']
        email = response.xpath('//input[@placeholder="Email Address"]/@value').get()

        if email:
            row_is.append(email)

        self.sheet.append(row_is)

        yield{
            "row": row_is
        }
        


    def spider_closed(self, spider):
        self.book.save(f'.//final sport men.xlsx')
  



        



process = CrawlerProcess()
process.crawl(EmailsSpider) #Here the name of the spider
process.start()
